"""
Genera data.js para la calculadora de flete de El Forastero.
Cruza Maestro_Estandarizado.xlsx + Lista_Precios_Estandarizada.xlsx
"""

import openpyxl
import json
import re
import os

# Rutas a los archivos fuente
BASE = "/Users/francoavacaperetti/Library/Application Support/Claude/local-agent-mode-sessions/90c57181-1e5e-4076-a655-9f37bc7fb1c0/ad610f33-4a84-4aaf-af17-a3d67859b83b/local_ab379a7d-7cb6-400d-b2ef-e5ad8143956e/outputs"
MAESTRO = os.path.join(BASE, "Maestro_Estandarizado.xlsx")
LISTA = os.path.join(BASE, "Lista_Precios_Estandarizada.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "data.js")


def parse_decimal(val):
    """Convierte string con coma decimal a float.
    Soporta: float/int nativo, '7.5' (punto decimal), '7,5' (coma decimal),
    '1.234,56' (miles con punto, decimal con coma).
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    # Si tiene coma: formato ARS (ej: "1.234,56" o "7,5")
    if ',' in s:
        s = s.replace(".", "").replace(",", ".")
    # Si solo tiene punto: es separador decimal directo (ej: "7.5")
    # → no tocar, parsear directo
    try:
        return float(s)
    except ValueError:
        return 0.0


def extract_pack_qty(um_pred):
    """Extrae cantidad de unidades del UM Pred (ej: PACKx10 -> 10, BOLx5 -> 5)."""
    if not um_pred:
        return 1
    m = re.search(r'x(\d+)', str(um_pred), re.IGNORECASE)
    return int(m.group(1)) if m else 1


def clean_rubro(rubro):
    """Quita el código entre paréntesis del rubro."""
    if not rubro:
        return ""
    return re.sub(r'\s*\(\d+\)\s*$', '', str(rubro)).strip()


def main():
    # Leer Maestro: codigo -> peso_kg
    wb_m = openpyxl.load_workbook(MAESTRO)
    ws_m = wb_m.active
    maestro = {}
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        cod = str(row[0]).lstrip('0')
        peso = parse_decimal(row[3])
        maestro[cod] = peso

    # Leer Lista de Precios
    wb_l = openpyxl.load_workbook(LISTA)
    ws_l = wb_l.active

    productos = []
    rubros_set = set()

    for row in ws_l.iter_rows(min_row=2, values_only=True):
        cod = str(row[0]).lstrip('0')
        descripcion = str(row[1]) if row[1] else ""
        rubro = clean_rubro(row[2])
        um_precio = str(row[3]) if row[3] else "unidad"
        precio_unit_iva = parse_decimal(row[6])    # Precio Venta con IVA (unidad)
        um_pred = str(row[7]) if row[7] else ""
        precio_pred_iva = parse_decimal(row[10])   # Precio Pred Venta con IVA

        if cod not in maestro:
            continue

        peso_unitario = maestro[cod]
        pack_qty = extract_pack_qty(um_pred)

        rubros_set.add(rubro)

        productos.append({
            "codigo": cod,
            "descripcion": descripcion,
            "rubro": rubro,
            "peso_unit_kg": round(peso_unitario, 4),
            "precio_unit_iva": round(precio_unit_iva, 2),
            "um_pred": um_pred if um_pred else "unidad",
            "pack_qty": pack_qty,
            "precio_pred_iva": round(precio_pred_iva, 2),
        })

    # Ordenar por rubro y descripcion
    productos.sort(key=lambda p: (p["rubro"], p["descripcion"]))
    rubros = sorted(rubros_set)

    # Tarifas de flete (flat, indexadas por lista)
    tarifas = {
        "Bari A":       {"tipo": "peso",       "valor": 75.0,  "zona": "Bariloche"},
        "Bari B":       {"tipo": "peso",       "valor": 230.0, "zona": "Bariloche"},
        "Cerv A":       {"tipo": "peso",       "valor": 75.0,  "zona": "Cervantes"},
        "Cerv B":       {"tipo": "peso",       "valor": 150.0, "zona": "Cervantes"},
        "Cerv C":       {"tipo": "peso",       "valor": 275.0, "zona": "Cervantes"},
        "Santa Rosa A": {"tipo": "porcentaje", "valor": 0.03,  "zona": "Santa Rosa"},
        "Santa Rosa B": {"tipo": "porcentaje", "valor": 0.05,  "zona": "Santa Rosa"},
        "Santa Rosa C": {"tipo": "porcentaje", "valor": 0.07,  "zona": "Santa Rosa"},
    }

    # Localidades con su lista asignada
    localidades = [
        {"nombre": "Bariloche",                    "cp": "8400", "provincia": "Río Negro",  "lista": "Bari A"},
        {"nombre": "Dina Huapi",                   "cp": "8402", "provincia": "Río Negro",  "lista": "Bari A"},
        {"nombre": "El Maitén",                    "cp": "9210", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "Esquel",                       "cp": "9200", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "El Hoyo",                      "cp": "8431", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "Lago Puelo",                   "cp": "9211", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "Epuyén",                       "cp": "9211", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "Trevelin",                     "cp": "9203", "provincia": "Chubut",     "lista": "Bari B"},
        {"nombre": "San Martín de los Andes",      "cp": "8370", "provincia": "Neuquén",    "lista": "Bari B"},
        {"nombre": "Junín de los Andes",           "cp": "8371", "provincia": "Neuquén",    "lista": "Bari B"},
        {"nombre": "Villa La Angostura",           "cp": "8407", "provincia": "Neuquén",    "lista": "Bari B"},
        {"nombre": "Villa Traful",                 "cp": "8403", "provincia": "Neuquén",    "lista": "Bari B"},
        {"nombre": "El Bolsón",                    "cp": "8430", "provincia": "Río Negro",  "lista": "Bari B"},
        {"nombre": "El Foyel",                     "cp": "8401", "provincia": "Río Negro",  "lista": "Bari B"},
        {"nombre": "El Manso",                     "cp": "8430", "provincia": "Río Negro",  "lista": "Bari B"},
        {"nombre": "Villa Llanquín",               "cp": "8401", "provincia": "Río Negro",  "lista": "Bari B"},
        {"nombre": "Villa Mascardi",               "cp": "8401", "provincia": "Río Negro",  "lista": "Bari B"},
        {"nombre": "Neuquén",                      "cp": "8300", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Plottier",                     "cp": "8316", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Cutral Có",                    "cp": "8322", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Centenario",                   "cp": "8309", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Plaza Huincul",                "cp": "8318", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Senillosa",                    "cp": "8320", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "San Patricio del Chañar",      "cp": "8305", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Vista Alegre Norte",           "cp": "8309", "provincia": "Neuquén",    "lista": "Cerv A"},
        {"nombre": "Cipolletti",                   "cp": "8324", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "General Roca",                 "cp": "8332", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Allen",                        "cp": "8328", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Cinco Saltos",                 "cp": "8303", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Villa Regina",                 "cp": "8336", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Ingeniero Huergo",             "cp": "8334", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Cervantes",                    "cp": "8326", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Chichinales",                  "cp": "8326", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Barda del Medio",              "cp": "8305", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Fernández Oro",                "cp": "8325", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Mainqué",                      "cp": "8326", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Campo Grande - San Isidro",    "cp": "8305", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Sargento Vidal",               "cp": "8305", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Campo Grande - Villa Manzano", "cp": "8308", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "General E. Godoy",             "cp": "8336", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Villa Manzano",                "cp": "8308", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Comandante Cordero",           "cp": "8301", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Valle Azul",                   "cp": "8336", "provincia": "Río Negro",  "lista": "Cerv A"},
        {"nombre": "Arroyito",                     "cp": "8313", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Piedra del Águila",            "cp": "8315", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Zapala",                       "cp": "8340", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Las Lajas",                    "cp": "8347", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Picún Leufú",                  "cp": "8313", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Villa El Chocón",              "cp": "8311", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Añelo",                        "cp": "8305", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Choele Choel",                 "cp": "8360", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Chimpay",                      "cp": "8364", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Catriel",                      "cp": "8307", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Luis Beltrán",                 "cp": "8361", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Coronel Belisle",              "cp": "8364", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Lamarque",                     "cp": "8363", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Chelforo",                     "cp": "8366", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Darwin",                       "cp": "8364", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "Pomona",                       "cp": "8363", "provincia": "Río Negro",  "lista": "Cerv B"},
        {"nombre": "25 de Mayo",                   "cp": "8201", "provincia": "La Pampa",   "lista": "Cerv B"},
        {"nombre": "Chos Malal",                   "cp": "8353", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Andacollo",                    "cp": "8353", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Loncopué",                     "cp": "8349", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Copahue",                      "cp": "8349", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Buta Ranquil",                 "cp": "8353", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Las Ovejas",                   "cp": "8353", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Bajada del Agrio",             "cp": "8353", "provincia": "Neuquén",    "lista": "Cerv B"},
        {"nombre": "Rincón de los Sauces",         "cp": "8319", "provincia": "Neuquén",    "lista": "Cerv C"},
        {"nombre": "Aluminé",                      "cp": "8345", "provincia": "Neuquén",    "lista": "Cerv C"},
        {"nombre": "Villa Pehuenia",               "cp": "8345", "provincia": "Neuquén",    "lista": "Cerv C"},
        {"nombre": "Santa Rosa",                   "cp": "6300", "provincia": "La Pampa",   "lista": "Santa Rosa A"},
        {"nombre": "Toay",                         "cp": "6303", "provincia": "La Pampa",   "lista": "Santa Rosa A"},
        {"nombre": "General Pico",                 "cp": "6360", "provincia": "La Pampa",   "lista": "Santa Rosa B"},
        {"nombre": "General Acha",                 "cp": "8200", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Eduardo Castex",               "cp": "6380", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Macachín",                     "cp": "6307", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Realicó",                      "cp": "6200", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Miguel Riglos",                "cp": "6301", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Bernardo Larroudé",            "cp": "6220", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Ingeniero Luiggi",             "cp": "6205", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Alta Italia",                  "cp": "6207", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Colonia Barón",                "cp": "6315", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Jacinto Aráuz",                "cp": "8208", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Intendente Alvear",            "cp": "6221", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Quemú Quemú",                  "cp": "6333", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Guatraché",                    "cp": "6311", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Bernasconi",                   "cp": "8204", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Parera",                       "cp": "6213", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Catriló",                      "cp": "6330", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Trenel",                       "cp": "6369", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Santa Isabel",                 "cp": "6323", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "General San Martín",           "cp": "8206", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Caleufú",                      "cp": "6387", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Alpachiri",                    "cp": "6309", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Victorica",                    "cp": "6319", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
        {"nombre": "Rivera",                       "cp": "6441", "provincia": "La Pampa",   "lista": "Santa Rosa C"},
    ]

    # Generar data.js
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write("// Generado automáticamente por generar_datos.py\n")
        f.write("// NO EDITAR MANUALMENTE\n\n")
        f.write(f"const PRODUCTOS = {json.dumps(productos, ensure_ascii=False, indent=2)};\n\n")
        f.write(f"const RUBROS = {json.dumps(rubros, ensure_ascii=False, indent=2)};\n\n")
        f.write(f"const LOCALIDADES = {json.dumps(localidades, ensure_ascii=False, indent=2)};\n\n")
        f.write(f"const TARIFAS = {json.dumps(tarifas, ensure_ascii=False, indent=2)};\n")

    print(f"Generado: {OUTPUT}")
    print(f"Productos: {len(productos)}")
    print(f"Rubros: {len(rubros)}")


if __name__ == "__main__":
    main()
